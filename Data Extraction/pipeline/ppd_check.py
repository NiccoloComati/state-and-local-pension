"""Redundant safety verifier: cross-check an extracted count total against the
Public Plans Database. The AV-independent second opinion - it catches whole
tables dropped or double-counted (which a within-table totals-check cannot see,
since a column shift conserves the total), and it works even for plans with no
human workbook. Free and offline.

For an active-member count distribution (Age_Serv_Num), the sum of the derived
grid should equal PPD `actives_tot` for that plan-year: verified exact on phx
(7941), mil (10974), aus (10149). We flag deviations beyond a tolerance.
"""
import functools
import os

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
PPD_PATH = os.path.join(ROOT, "Data", "Common", "states", "ppd-data-latest.xlsx")


@functools.lru_cache(maxsize=1)
def _load(fy=2019):
    """{ppd_id: actives_tot} for the given fiscal year, or {} if the PPD file
    is not present (e.g. not uploaded to the cluster) - callers degrade
    gracefully."""
    if not os.path.exists(PPD_PATH):
        return {}
    import openpyxl
    wb = openpyxl.load_workbook(PPD_PATH, read_only=True, data_only=True)
    try:
        ws = wb.active
        hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
        idx = {h: i for i, h in enumerate(hdr)}
        need = ("ppd_id", "fy", "actives_tot")
        if not all(k in idx for k in need):
            return {}
        out = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[idx["fy"]] == fy:
                pid, a = r[idx["ppd_id"]], r[idx["actives_tot"]]
                if isinstance(pid, (int, float)) and isinstance(a, (int, float)):
                    out[int(pid)] = a
        return out
    finally:
        wb.close()


def actives_tot(ppd_id, fy=2019):
    """PPD active-member total for a plan-year, or None if unavailable."""
    if ppd_id is None:
        return None
    return _load(fy).get(int(ppd_id))


def cross_check(derived, ppd_id, tol=0.02, fy=2019):
    """Compare the sum of a derived COUNT grid to PPD actives_tot. Returns a
    dict {expected, extracted, ratio, status} or None if no PPD figure exists.
    status: 'ok' within tol, 'off' beyond it, 'no_ppd' when the base is 0/None."""
    expected = actives_tot(ppd_id, fy)
    if not expected:
        return None
    extracted = sum(v for row in derived.get("cells", [])
                    for v in row if isinstance(v, (int, float)) and not isinstance(v, bool))
    ratio = extracted / expected if expected else None
    status = "ok" if ratio is not None and abs(ratio - 1.0) <= tol else "off"
    return {"expected": expected, "extracted": extracted,
            "ratio": round(ratio, 4) if ratio is not None else None, "status": status}
