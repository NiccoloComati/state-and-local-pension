"""
Provenance catalogue generator (2026-06-11). Companion to
`Documentation/model_input_dictionary.md` (the schema side).

Outputs (all into Documentation/):
  - state_sheet_fill_audit.csv : value-signature audit of the 40 state plan
    workbooks ([PLAN]_2017.xlsx, 9 sheets) — same method as the city scan:
    plan_specific / shared_default (identical numeric content across >1 plan)
    / empty / absent.
  - state_notes_harvest.md : verbatim dump of every state workbook's notes
    sheet(s) + richness classification (the only in-repo Brookings provenance).
  - provenance_register.csv : the combined states+cities register — one row per
    plan x input element, with source channel, vintage, specificity, extractor,
    evidence, confidence. Tier-A columns are mechanical; Tier-B evidence comes
    from notes sheets / csv_matrices / city logs; Tier-C gaps are recorded as
    explicit 'undocumented' rather than silently omitted.

Inputs it reads (read-only):
  - Data/Plans/States/[PLAN]/[PLAN]_2017.xlsx           (40 workbooks)
  - Code/python/fast/Main_PensionModel.py               (AVAILABLE_DATA parsed by regex)
  - Data/Common/states/ppd-data-latest.xlsx             (fy2022 scalar resolution)
  - Data/Common/states/PPD_planlevel_main.csv           (legacy 2017 fallbacks)
  - Data/Common/states/PPD_planlevel_main_updated.csv   (pctmale/pctmrg/reduct/inactive_adj)
  - Data/Common/states/planchanges_main_2022_clean.xlsx (tier-rule coverage)
  - Data/Sources/brookings_package_csv_matrices/        (granular extraction layer; the FULLER copy)
  - Documentation/city_sheet_fill_audit.csv + city_source_inventory.csv
    (run Documentation/city_data_scan.py first if missing)

Run:  python Documentation/provenance_scan.py   (from the project root)
"""
import os, re, csv, glob, hashlib
from collections import defaultdict

import openpyxl
import pandas as pd
import numpy as np

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOCS = os.path.join(ROOT, "Documentation")
STATES = os.path.join(ROOT, "Data", "Plans", "States")
COMMON = os.path.join(ROOT, "Data", "Common", "states")
MATRICES = os.path.join(ROOT, "Data", "Sources", "brookings_package_csv_matrices")
FAST_RUNNER = os.path.join(ROOT, "Code", "python", "fast", "Main_PensionModel.py")

STATE_SHEETS = ["ageservice", "retdist", "wagerel", "mortality", "wagegrowth",
                "withdrawal", "retirement", "refund", "disability"]
GHOST_SHEETS = {"wagegrowth": "engine uses PPD wage-growth chain instead",
                "disability": "engine uses constant DisabilityPayoutRate=0.025"}
# city collection sheet -> state model input
CITY_TO_STATE = {"Age_Serv_Num": "ageservice", "Retirement": "retdist",
                 "Age_Serv_Wage": "wagerel", "Avg_Mort": "mortality",
                 "Wage_Growth": "wagegrowth", "Sep_Rate": "withdrawal",
                 "Ret_Rate": "retirement", "Refund_Rate": "refund",
                 "Inactv_Serv_Num": "inactive_distribution"}
# state sheet -> csv_matrices precursor file prefixes
MATRIX_PRECURSORS = {"ageservice": ["ageservice"], "retdist": ["retdist", "retbenrel"],
                     "wagerel": ["wagerel"], "mortality": ["mortality"],
                     "wagegrowth": ["wagegrowth"], "withdrawal": ["withdrawal"],
                     "retirement": ["retirement"], "refund": ["refund"],
                     "disability": ["disability"]}


def parse_available_data():
    """Parse the AVAILABLE_DATA dict out of the fast runner source (regex, no import
    — the module runs argparse at import time)."""
    txt = open(FAST_RUNNER, encoding="utf-8").read()
    block = txt.split("AVAILABLE_DATA = {")[1].split("}")[0]
    out = {}
    for m in re.finditer(r"'(\w+)':\s*\[([^\]]+)\]", block):
        plan, flags = m.group(1), m.group(2)
        out[plan] = [f.strip() == "True" for f in flags.split(",")]
    return out


def sheet_sig(ws):
    vals = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                vals.append(round(float(v), 4))
    if not vals:
        return None, 0
    return hashlib.md5(repr(vals).encode()).hexdigest()[:8], len(vals)


def harvest_notes(wb):
    """Return (verbatim_lines, classification) from any notes-ish sheet."""
    lines = []
    for sn in wb.sheetnames:
        if "note" in sn.lower() or "scratch" in sn.lower():
            for row in wb[sn].iter_rows(values_only=True):
                txt = " | ".join(str(v).strip() for v in row if v is not None and str(v).strip())
                if txt:
                    lines.append(txt)
    blob = " ".join(lines).lower()
    has_url = "http" in blob or "www." in blob
    has_author = "author" in blob or "brookings" in blob
    cls = ("rich_with_source_urls" if has_url else
           "rich" if (has_author or len(lines) > 8) else
           "thin" if lines else "none")
    return lines, cls


def first_nonmissing(*vals):
    for v in vals:
        try:
            f = float(v)
            if not np.isnan(f):
                return f
        except (TypeError, ValueError):
            continue
    return float("nan")


def main():
    avail = parse_available_data()           # 37 fast-runner plans
    all_plans = sorted(os.path.basename(d) for d in glob.glob(os.path.join(STATES, "*"))
                       if os.path.isdir(d))

    # ---- load common scalar sources once ----
    ppd = pd.read_excel(os.path.join(COMMON, "ppd-data-latest.xlsx"),
                        sheet_name="ppd-data-latest", header=0)
    legacy17 = pd.read_csv(os.path.join(COMMON, "PPD_planlevel_main.csv"))
    legacy22 = pd.read_csv(os.path.join(COMMON, "PPD_planlevel_main_updated.csv"))
    tier_in = pd.read_excel(os.path.join(COMMON, "planchanges_main_2022_clean.xlsx"),
                            sheet_name="in", header=0)
    tier_planids = set(tier_in["planid"].astype(str))

    def leg17(plan, col):
        row = legacy17[legacy17["planid"] == f"{plan}_2017"]
        return first_nonmissing(row[col].iloc[0]) if len(row) and col in row.columns else float("nan")

    def leg22(plan, col):
        row = legacy22[legacy22["planid"] == f"{plan}_2022"]
        return first_nonmissing(row[col].iloc[0]) if len(row) and col in row.columns else float("nan")

    # ---- pass over state workbooks: signatures + notes ----
    sigs = {}                                    # plan -> {sheet: (sig, n)}
    bysheet = defaultdict(lambda: defaultdict(list))
    notes_all = {}                               # plan -> (lines, cls)
    for plan in all_plans:
        f = os.path.join(STATES, plan, f"{plan}_2017.xlsx")
        if not os.path.exists(f):
            continue
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        sm = {s.lower(): s for s in wb.sheetnames}
        sigs[plan] = {}
        for sheet in STATE_SHEETS:
            if sheet in sm:
                sig, n = sheet_sig(wb[sm[sheet]])
                sigs[plan][sheet] = (sig, n)
                bysheet[sheet][sig if sig else "EMPTY"].append(plan)
            else:
                sigs[plan][sheet] = ("ABSENT", 0)
                bysheet[sheet]["ABSENT"].append(plan)
        notes_all[plan] = harvest_notes(wb)
        wb.close()

    def specificity(sheet, sig):
        if sig == "ABSENT": return "absent"
        if sig in (None, "EMPTY"): return "empty"
        return "shared_default" if len(bysheet[sheet][sig]) > 1 else "plan_specific"

    # ---- output 1: state sheet fill audit ----
    with open(os.path.join(DOCS, "state_sheet_fill_audit.csv"), "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["plan", "sheet", "n_numeric_cells", "signature", "status",
                    "n_plans_sharing_signature", "availableData_flag"])
        for plan in sorted(sigs):
            flags = avail.get(plan)
            for i, sheet in enumerate(STATE_SHEETS):
                sig, n = sigs[plan][sheet]
                st = specificity(sheet, "EMPTY" if sig is None else sig)
                share = len(bysheet[sheet][sig]) if sig not in ("ABSENT", None, "EMPTY") else ""
                flag = (flags[i] if flags else "")
                w.writerow([plan, sheet, n, sig if sig and sig != "ABSENT" else "",
                            st, share, flag])

    # ---- output 2: notes harvest ----
    with open(os.path.join(DOCS, "state_notes_harvest.md"), "w", encoding="utf-8") as fh:
        fh.write("# State Workbook Notes Harvest\n\n")
        fh.write("Verbatim contents of every `notes`-type sheet in the 40 state plan "
                 "workbooks — the only in-repo trace of Brookings' extraction provenance. "
                 "Generated by `provenance_scan.py`; do not hand-edit.\n\n")
        counts = defaultdict(int)
        for plan in sorted(notes_all):
            counts[notes_all[plan][1]] += 1
        fh.write("**Summary:** " + ", ".join(f"{k}: {v}" for k, v in sorted(counts.items())) + "\n\n")
        fh.write("| classification | meaning |\n|---|---|\n"
                 "| rich_with_source_urls | names the exact AV/CAFR documents (URLs) |\n"
                 "| rich | author/date/plan metadata or substantive notes, no URLs |\n"
                 "| thin | trivial content (e.g. just a title) |\n| none | no notes sheet |\n\n")
        for plan in sorted(notes_all):
            lines, cls = notes_all[plan]
            fh.write(f"## {plan}  —  `{cls}`\n\n")
            if lines:
                for ln in lines:
                    fh.write(f"    {ln}\n")
            else:
                fh.write("    (no notes sheet)\n")
            fh.write("\n")

    # ---- csv_matrices granular-layer presence per plan ----
    def matrices_for(plan):
        for cand in (f"{plan}_2017", plan):
            d = os.path.join(MATRICES, cand)
            if os.path.isdir(d):
                return sorted(os.path.splitext(b)[0] for b in os.listdir(d) if b.endswith(".csv"))
        return []

    # ---- scalar resolution (emulates the engine's fallback helpers) ----
    def resolve_scalars(plan):
        prow = ppd[(ppd["ppd_id"] == int("".join(filter(str.isdigit, plan)))) & (ppd["fy"] == 2022)]
        if not len(prow):
            return {}
        g = lambda c: first_nonmissing(prow[c].iloc[0]) if c in prow.columns else float("nan")
        out = {}
        # wage growth chain
        chain = [("ppd_panel.PayrollGrowthAssumption", g("PayrollGrowthAssumption")),
                 ("ppd_panel.WageInflation", g("WageInflation")),
                 ("legacy_ppd_csv.wage_inf[2017]", leg17(plan, "wage_inf")),
                 ("ppd_panel.InflationAssumption_GASB", g("InflationAssumption_GASB")),
                 ("legacy_ppd_csv.inflation[2017]", leg17(plan, "inflation"))]
        out["WageGrowth"] = next((name for name, v in chain if not np.isnan(v)), "UNRESOLVED")
        # inflation chain
        chain = [("ppd_panel.InflationAssumption_GASB", g("InflationAssumption_GASB")),
                 ("legacy_ppd_csv.inflation[2017]", leg17(plan, "inflation"))]
        out["Inflation"] = next((name for name, v in chain if not np.isnan(v)), "UNRESOLVED")
        # inactive scaling
        adj = leg22(plan, "inactive_adj")
        if np.isnan(adj):
            out["InactiveScaling"] = "UNRESOLVED (inactive_adj missing)"
        elif adj == 1.0:
            chain = [("ppd_panel.InactiveVestedMembers", g("InactiveVestedMembers")),
                     ("legacy_ppd_csv.inactive[2017]", leg17(plan, "inactive"))]
            out["InactiveScaling"] = next((name for name, v in chain if not np.isnan(v)), "UNRESOLVED")
        else:
            out["InactiveScaling"] = f"ppd_panel.actives_tot x inactive_adj({adj:g})"
        return out

    # ---- output 3: the combined register ----
    reg_path = os.path.join(DOCS, "provenance_register.csv")
    with open(reg_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["track", "plan", "element", "element_kind", "consumed_by_engine",
                    "source_channel", "source_file", "vintage", "specificity",
                    "extractor", "evidence", "confidence", "note"])

        # ----- STATE rows -----
        for plan in sorted(sigs):
            flags = avail.get(plan)
            model_status = ("fast_runner" if plan in avail else
                            "r_track_only" if plan == "MA50" else "no_2022_track")
            notes_cls = notes_all[plan][1]
            mats = matrices_for(plan)
            for i, sheet in enumerate(STATE_SHEETS):
                sig, n = sigs[plan][sheet]
                spec = specificity(sheet, "EMPTY" if sig is None else sig)
                pre = [m for m in mats if any(m.startswith(p) for p in MATRIX_PRECURSORS[sheet])]
                ev = []
                if notes_cls != "none": ev.append(f"workbook notes: {notes_cls}")
                if pre: ev.append(f"csv_matrices precursors: {'+'.join(pre)}")
                ev.append("source AV/CAFR PDFs in plan folder")
                if sheet in GHOST_SHEETS:
                    consumed, channel, conf = "no_ghost_sheet", "n/a", "high"
                    note = GHOST_SHEETS[sheet]
                elif flags is None:
                    consumed, conf, note = model_status, "high", "plan not in fast runner"
                    channel = "plan_workbook"
                elif flags[i]:
                    consumed, channel, conf = "yes", "plan_workbook", "high"
                    note = ""
                else:
                    consumed, channel, conf = "fallback", "default_assumptions", "high"
                    note = "availableData=False"
                w.writerow(["state", plan, sheet, "distribution_sheet", consumed,
                            channel,
                            f"Data/Plans/States/{plan}/{plan}_2017.xlsx" if channel == "plan_workbook"
                            else ("Data/Common/states/default_assumptions.xlsx" if channel == "default_assumptions" else ""),
                            "FY2017", spec, "brookings", "; ".join(ev), conf,
                            note if note else ("extraction assumptions undocumented (Tier C)"
                                               if notes_cls in ("thin", "none") and channel == "plan_workbook" else "")])
            # scalar resolutions (2022-track plans only)
            if plan in avail or plan == "MA50":
                for elem, src in resolve_scalars(plan).items():
                    w.writerow(["state", plan, elem, "scalar_resolution", "yes",
                                src.split(".")[0] if "." in src else "ppd_panel",
                                "Data/Common/states/ppd-data-latest.xlsx", "fy2022",
                                "plan_specific", "CRR_digest", f"resolved: {src}",
                                "high", "fallback chain per model_input_dictionary.md §3"])
            # tier rules
            has_tier = f"{plan}_2022" in tier_planids
            w.writerow(["state", plan, "tier_rules", "tier_rules",
                        "yes" if has_tier else "missing",
                        "planchanges", "Data/Common/states/planchanges_main_2022_clean.xlsx",
                        "2022 update (curated 2024)", "plan_specific" if has_tier else "absent",
                        "in-house (from Brookings planchanges_main)",
                        "row present in sheet 'in'" if has_tier else "no planid row",
                        "high", ""])
            # demographics
            has_demo = len(legacy22[legacy22["planid"] == f"{plan}_2022"]) > 0
            w.writerow(["state", plan, "pctmale/pctmrg/reduct/inactive_adj", "scalar_resolution",
                        "yes" if has_demo else "missing", "legacy_ppd_csv",
                        "Data/Common/states/PPD_planlevel_main_updated.csv", "2022 file",
                        "plan_specific" if has_demo else "absent", "brookings-era",
                        "planid row present" if has_demo else "no row", "high",
                        "underlying derivation undocumented (Tier C)"])

        # ----- CITY rows (from the city scan CSVs) -----
        fill = pd.read_csv(os.path.join(DOCS, "city_sheet_fill_audit.csv"))
        inv = pd.read_csv(os.path.join(DOCS, "city_source_inventory.csv")).set_index("plan")
        for _, r in fill.iterrows():
            plan = r["plan"]
            meta = inv.loc[plan] if plan in inv.index else None
            year = re.search(r"data(\d\d)", plan)
            vintage = f"FY20{year.group(1)}" if year else "FY?"
            ev = []
            if meta is not None:
                if meta["city_logs"]: ev.append(f"{int(meta['city_logs'])} log file(s) in city folder")
                if meta["city_AV_pdfs"]: ev.append(f"{int(meta['city_AV_pdfs'])} AV PDF(s) in folder")
                if isinstance(meta["city_ppd_ids"], str) and meta["city_ppd_ids"]:
                    ev.append(f"city ppd_ids {meta['city_ppd_ids']}")
            w.writerow(["city", plan, f"{r['sheet']} -> {CITY_TO_STATE[r['sheet']]}",
                        "distribution_sheet", "not_yet_integrated", "plan_workbook",
                        f"Data/Plans/Cities/{plan.split('_')[0]}_modeldata/{plan}.xlsx",
                        vintage, r["status"], "in-house (Fan/Gant 2022)",
                        "; ".join(ev) if ev else "no in-folder provenance",
                        "high" if r["status"] in ("plan_specific", "empty", "absent") else "medium",
                        "Airtable + _log.md carry assumptions" if ev else
                        "provenance only in Airtable (re-export pending)"])
        for plan in inv.index:
            if str(plan).startswith("(none"):
                continue
            tv = inv.loc[plan, "city_tiervars"]
            w.writerow(["city", plan, "tier_rules", "tier_rules", "not_yet_integrated",
                        "tiervars_workbook" if tv else "missing",
                        f"Data/Plans/Cities/{str(plan).split('_')[0]}_modeldata/ (tiervars)" if tv else "",
                        "FY2019-ish", "plan_specific" if tv else "absent",
                        "in-house (Fan/Gant 2022)",
                        f"{int(tv)} tiervars workbook(s) in city folder" if tv else "none",
                        "high", "needs reshaping to planchanges schema (city_data_audit §6.1)"])

    n = sum(1 for _ in open(reg_path, encoding="utf-8")) - 1
    print(f"wrote state_sheet_fill_audit.csv, state_notes_harvest.md, provenance_register.csv ({n} rows)")


if __name__ == "__main__":
    main()
