"""
City-data audit scan (2026-06-11). Reproducible evidence behind
`city_data_audit.md` §3.1 and `data_sources_map.md`.

Emits two CSVs into Documentation/:
  - city_sheet_fill_audit.csv : per plan-workbook x canonical sheet, the numeric
    cell count + value-signature + status (plan_specific / shared_default /
    empty / absent). "shared_default" = the sheet's numeric content is byte-for-
    byte identical across >1 plan, i.e. a copied default table, not a real
    extraction.
  - city_source_inventory.csv : per plan-workbook, the folder/source provenance
    (city, collection generation, city-level ppd_ids, AV/CAFR/tiervars/log
    presence).

Run:  python Documentation/city_data_scan.py
(from the project root). Read-only; writes only the two CSVs.
"""
import os, re, glob, csv, hashlib
from collections import defaultdict

ROOT   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CITIES = os.path.join(ROOT, "Data", "Plans", "Cities")
DOCS   = os.path.join(ROOT, "Documentation")

import openpyxl

CANON = ["Wage_Growth", "Refund_Rate", "Avg_Mort", "Sep_Rate", "Ret_Rate",
         "Retirement", "Age_Serv_Num", "Age_Serv_Wage", "Inactv_Serv_Num"]
def norm(s): return re.sub(r'[^a-z]', '', s.lower())
CANON_N = {norm(c): c for c in CANON}

# Cities whose folder uses the single "primary"+"tier" workbook layout (no per-fund
# split, no tiervars/logs/PDFs) rather than one-workbook-per-fund.
PRIMARY_TIER_LAYOUT = {"den", "fw", "nsh", "nyc", "sea", "aus", "ind"}

def is_plan_wb(fn):
    b = os.path.basename(fn).lower()
    if any(x in b for x in ["tiervars", "_tier", "tier1", "tier2", "tier3", "tier4",
                            "tier5", "tier6", "tier7", "tier8", "tier9", "working",
                            "template", "overview", "tiersasy", "variablesdb",
                            "default_assump", "migration"]):
        return False
    return ("_data" in b) and b.endswith(".xlsx")

def sheet_sig(ws):
    vals = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                vals.append(round(float(v), 4))
    if not vals:
        return None, 0
    return hashlib.md5(repr(vals).encode()).hexdigest()[:8], len(vals)

# ---- pass 1: collect per-plan per-sheet (sig, ncells) ----
wbs = sorted(f for f in glob.glob(os.path.join(CITIES, "*", "*.xlsx")) if is_plan_wb(f))
rows = {}                                   # plan -> {sheet: (sig, n)}
bysheet = defaultdict(lambda: defaultdict(list))   # sheet -> sig -> [plans]
for f in wbs:
    plan = os.path.basename(f)[:-5]
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    sm = {norm(s): s for s in wb.sheetnames}
    rows[plan] = {}
    for cn, canon in CANON_N.items():
        if cn in sm:
            sig, n = sheet_sig(wb[sm[cn]])
            rows[plan][canon] = (sig, n)
            key = sig if sig else "EMPTY"
            bysheet[canon][key].append(plan)
        else:
            rows[plan][canon] = ("ABSENT", 0)
            bysheet[canon]["ABSENT"].append(plan)
    wb.close()

def status_for(canon, sig):
    if sig in ("ABSENT", None):
        return "absent"
    if sig == "EMPTY":
        return "empty"
    return "shared_default" if len(bysheet[canon][sig]) > 1 else "plan_specific"

with open(os.path.join(DOCS, "city_sheet_fill_audit.csv"), "w", newline="", encoding="utf-8") as fh:
    w = csv.writer(fh)
    w.writerow(["plan", "city", "sheet", "n_numeric_cells", "signature",
                "status", "n_plans_sharing_signature"])
    for plan in sorted(rows):
        city = plan.split("_")[0]
        for canon in CANON:
            sig, n = rows[plan][canon]
            st = status_for(canon, "EMPTY" if (sig is None) else sig)
            share = (len(bysheet[canon][sig]) if sig not in ("ABSENT", None, "EMPTY") else "")
            w.writerow([plan, city, canon, n, (sig if sig else ""), st, share])

# ---- pass 2: source/provenance inventory ----
def folder_meta(d):
    av = cafr = tv = log = 0
    ppds = set()
    for root, _, files in os.walk(d):
        if "archive" in root.lower() and os.path.normcase(root) != os.path.normcase(d):
            continue
        for b in files:
            bl = b.lower()
            if bl.endswith(".pdf"):
                m = re.search(r'_(\d{2,3})\.pdf$', bl)
                if m: ppds.add(m.group(1))
                if "_av_" in bl or "av_2019" in bl or "_av2019" in bl or "merf_av" in bl:
                    av += 1
                elif any(k in bl for k in ("cafr", "acfr", "financialstatements", "annualreport")):
                    cafr += 1
            elif bl.endswith("tiervars.xlsx"): tv += 1
            elif bl.endswith(".md") or "_log" in bl or bl.startswith("log") or "updatelog" in bl:
                log += 1
    return sorted(ppds), av, cafr, tv, log

city_meta = {}
for d in sorted(os.listdir(CITIES)):
    full = os.path.join(CITIES, d)
    if os.path.isdir(full) and d != "_migration":
        city_meta[d.split("_")[0]] = folder_meta(full)

with open(os.path.join(DOCS, "city_source_inventory.csv"), "w", newline="", encoding="utf-8") as fh:
    w = csv.writer(fh)
    w.writerow(["plan", "city", "folder_layout", "city_ppd_ids", "city_AV_pdfs",
                "city_CAFR_pdfs", "city_tiervars", "city_logs"])
    seen_cities = set()
    for plan in sorted(rows):
        city = plan.split("_")[0]
        layout = "primary+tier" if city in PRIMARY_TIER_LAYOUT else "per-fund"
        ppds, av, cafr, tv, log = city_meta.get(city, ([], 0, 0, 0, 0))
        w.writerow([plan, city, layout, "|".join(ppds), av, cafr, tv, log])
        seen_cities.add(city)
    # cities with folders but no scannable plan workbook (empty/placeholder)
    for city, (ppds, av, cafr, tv, log) in sorted(city_meta.items()):
        if city not in seen_cities:
            w.writerow([f"(none: {city})", city, "empty", "|".join(ppds),
                        av, cafr, tv, log])

print("wrote city_sheet_fill_audit.csv and city_source_inventory.csv to Documentation/")
